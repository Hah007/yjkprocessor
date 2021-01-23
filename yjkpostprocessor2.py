#! python3
#-*- coding: gbk -*-

#针对盈建科1.9.0(2018.05.18)版本编程
#目前仅处理圆钢管混凝土柱
#concrete-filled circular steel tube frame columns
#   R1:表示圆钢管混凝土柱轴压力与轴心受压承载力的比值F1/f
#   R2:表示圆钢管混凝土柱拉弯受力时外力与承载力的比值F2/f
#   R3:表示圆钢管混凝土柱设计剪力与受剪承载力的比值F3/f

#有机会再加上处理：方钢管混凝土柱和剪力墙

import os
import openpyxl
import re

#   构件编号
ncngnumber = []
#   构件类型
coltype = []
#   外部直径
outdiameter = []
#   内部直径
indiameter = []
#   轴压比
uc = []
#   F1
rf1 = []
rf1gamma = []
#   F2
rf2 = []
rf2gamma = []
#   F3
rf3 = []
rf3gamma = []
#   层数
nfloor = []

def readmass():
    thefile = ".\设计结果\wmass.out"
    
    if not os.path.isfile(thefile):
        print("wass.out文件不存在，或者没有计算过。")
    else:
        with open(thefile) as file_object:
            for line in file_object:
                #总的楼层组装数
                if line.endswith('楼层属性\n'):
                    for j in range(3):
                        next(file_object)
                    sumfloor = int(next(file_object).split()[0])
    return sumfloor

def readwpj():
    floornum = readmass()
    for i in range(1, floornum + 1):
        thefile = ".\设计结果\wpj" + str(i) + ".out"
        if not os.path.isfile(thefile):
            print("wpj "+ str(i) +" .out文件不存在，或者没有计算过。")
        else:
            wpjfile = open(thefile)
            wpjfileline = wpjfile.readlines()
            for j in range(len(wpjfileline)):
                
                #   处理圆钢管混凝土直柱
                if wpjfileline[j].endswith("圆钢管砼柱\n"):
                    
                    templine = wpjfileline[j - 2]
                    
                    #   构件编号
                    reline1 = re.compile(r'N-C=(\d+)')
                    notemp = reline1.search(templine)
                    ncngnumber.append(notemp.group(1))
                    
                    #   B 和 H
                    reline2 = re.compile(r'(\d+)\*(\d+)\*0\*0\*0\*0')
                    bhtemp = reline2.search(templine)
                    outdiameter.append(bhtemp.group(1))
                    indiameter.append(bhtemp.group(2))                  
                    
                    #   构件类型
                    coltype.append("圆钢管砼柱")

                    #   轴压比
                    #   圆钢管混凝土柱没有轴压比限值，直接取值
                    uc.append(wpjfileline[j + 3].split()[-1])
                    
                    #   F1,F2,F3还要有超限提示信息!!!
                    #   会变"<"和">"

                    #   F1
                    f1line = wpjfileline[j + 4]

                    f1reline = re.compile(r'R_F1=     (\d.\d\d\d)')
                    f1temp = f1reline.search(f1line)
                    rf1.append(f1temp.group(1))
                    
                    f1gammareline = re.compile(r'1/γ=     (\d.\d\d\d)')
                    f1temp = f1gammareline.search(f1line)
                    rf1gamma.append(f1temp.group(1))
                    
                    #   F2 和 F3
                    #   F2:《钢管砼规范》6.1.8
                    
                    f2line = wpjfileline[j + 5]
                    
                    f2reline = re.compile(r'R_F2=     (\d.\d\d\d)')
                    f2temp = f2reline.search(f2line)
                    
                    f23reline = re.compile(r'R_F3=     (\d.\d\d\d)')
                    f23temp = f23reline.search(f2line)
                    
                    f2gammareline = re.compile(r'1/γ=     (\d.\d\d\d)')
                    f2gammatemp = f2gammareline.search(f2line)
                    
                    if f2temp == None:
                        #   F2
                        rf2.append(" ")
                        rf2gamma.append(" ")
                        
                        #   F3
                        rf3.append(f23temp.group(1))
                        rf3gamma.append(f2gammatemp.group(1))
                        
                    else:
                        #   F2
                        rf2.append(f2temp.group(1))
                        rf2gamma.append(f2gammatemp.group(1))
                        
                        #   F3
                        f3line = wpjfileline[j + 6]
                        
                        f3reline = re.compile(r'R_F3=     (\d.\d\d\d)')
                        f3temp = f3reline.search(f3line)
                        rf3.append(f3temp.group(1))
                        
                        f3gammareline = re.compile(r'1/γ=     (\d.\d\d\d)')
                        f3gammatemp = f3gammareline.search(f3line)
                        rf3gamma.append(f3gammatemp.group(1))
                    
                    #   层数
                    nfloor.append(i)
                    
                #   处理圆钢管混凝土斜柱
                elif wpjfileline[j].endswith("圆钢管砼斜柱\n"):
                    
                    templine = wpjfileline[j - 2]
                    
                    #   构件编号
                    reline1 = re.compile(r'N-G=(\d+)')
                    notemp = reline1.search(templine)
                    ncngnumber.append(notemp.group(1))
                    
                    #   B 和 H
                    reline2 = re.compile(r'(\d+)\*(\d+)\*0\*0\*0\*0')
                    bhtemp = reline2.search(templine)
                    outdiameter.append(bhtemp.group(1))
                    indiameter.append(bhtemp.group(2))                  
                    
                    #   构件类型
                    coltype.append("圆钢管砼斜柱")

                    #   轴压比
                    #   圆钢管混凝土柱没有轴压比限值，直接取值
                    uc.append(wpjfileline[j + 3].split()[-1])
                    
                    #   F1,F2,F3还要有超限提示信息!!!
                    #   会变"<"和">"

                    #   F1
                    f1line = wpjfileline[j + 4]

                    f1reline = re.compile(r'R_F1=     (\d.\d\d\d)')
                    f1temp = f1reline.search(f1line)
                    rf1.append(f1temp.group(1))
                    
                    f1gammareline = re.compile(r'1/γ=     (\d.\d\d\d)')
                    f1temp = f1gammareline.search(f1line)
                    rf1gamma.append(f1temp.group(1))
                    
                    #   F2 和 F3
                    f2line = wpjfileline[j + 5]
                    
                    f2reline = re.compile(r'R_F2=     (\d.\d\d\d)')
                    f2temp = f2reline.search(f2line)
                    
                    f23reline = re.compile(r'R_F3=     (\d.\d\d\d)')
                    f23temp = f23reline.search(f2line)
                    
                    f2gammareline = re.compile(r'1/γ=     (\d.\d\d\d)')
                    f2gammatemp = f2gammareline.search(f2line)
                    
                    if f2temp == None:
                        #   F2
                        rf2.append("")
                        rf2gamma.append("")
                        
                        #   F3
                        rf3.append(f23temp.group(1))
                        rf3gamma.append(f2gammatemp.group(1))
                        
                    else:
                        #   F2
                        rf2.append(f2temp.group(1))
                        rf2gamma.append(f2gammatemp.group(1))
                        
                        #   F3
                        f3line = wpjfileline[j + 6]
                        
                        f3reline = re.compile(r'R_F3=     (\d.\d\d\d)')
                        f3temp = f3reline.search(f3line)
                        rf3.append(f3temp.group(1))
                        
                        f3gammareline = re.compile(r'1/γ=     (\d.\d\d\d)')
                        f3gammatemp = f3gammareline.search(f3line)
                        rf3gamma.append(f3gammatemp.group(1))
                    
                    #   层数
                    nfloor.append(i)
                        
def createexcel():
    wb = openpyxl.Workbook()
    sheetygtz = wb.create_sheet(index = 0, title = '圆钢管混凝土柱')
    sheetfgtz = wb.create_sheet(index = 1, title = '方钢管混凝土柱')
    sheetjlq = wb.create_sheet(index = 2, title = '剪力墙')
    
    #   N-C柱编号，N-G支撑编号
    sheetygtz['A1'] = 'N-C/N-G'
    sheetygtz['B1'] = '构件类型'
    sheetygtz['C1'] = 'B'
    sheetygtz['D1'] = 'H'
    sheetygtz['E1'] = 'Uc'
    sheetygtz['F1'] = 'R_F1'
    sheetygtz['G1'] = '1/γ'
    sheetygtz['H1'] = 'R_F2'
    sheetygtz['I1'] = '1/γ'
    sheetygtz['J1'] = 'R_F3'
    sheetygtz['K1'] = '1/γ'
    sheetygtz['L1'] = '层数'
    
    #   填写各个数据
    for row  in range(2, len(coltype) + 2):
        sheetygtz.cell(column = 1, row = row, value = int(ncngnumber[row - 2]))
        sheetygtz.cell(column = 2, row = row, value = coltype[row - 2])
        sheetygtz.cell(column = 3, row = row, value = int(outdiameter[row - 2]))
        sheetygtz.cell(column = 4, row = row, value = int(indiameter[row - 2]))
        sheetygtz.cell(column = 5, row = row, value = float(uc[row - 2]))
        sheetygtz.cell(column = 6, row = row, value = float(rf1[row - 2]))
        sheetygtz.cell(column = 7, row = row, value = float(rf1gamma[row - 2]))
        
        if rf2[row - 2] != ' ':
            sheetygtz.cell(column = 8, row = row, value = float(rf2[row - 2]))
            sheetygtz.cell(column = 9, row = row, value = float(rf2gamma[row - 2]))
        
        sheetygtz.cell(column = 10, row = row, value = float(rf3[row - 2]))
        sheetygtz.cell(column = 11, row = row, value = float(rf3gamma[row - 2]))
        sheetygtz.cell(column = 12, row = row, value = nfloor[row - 2])
    
    #   调整 B 列宽度
    sheetygtz.column_dimensions['B'].width = 12
    
    #   冻结首行
    sheetygtz.freeze_panes = 'A2'
    
    wb.save('盈建科计算结果读取软件(钢管混凝土柱、剪力墙)v1.xlsx')
    
if __name__ == '__main__':
    readwpj()
    createexcel()
