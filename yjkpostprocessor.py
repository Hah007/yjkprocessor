#! python3
#

# 针对盈建科1.8.3.0(2017.09.20)版本
# 
# 仅针对单塔模型
# 目前统计地下室相关信息，如果要不统计地下室信息，也可以做出来
# 目前仅做了地震位移角，风位移角是一样的，还没有做

import os
from fractions import Fraction
import openpyxl

#刚度比
ratx1 = []
raty1 = []

ratx2 = []
raty2 = []

#位移角

#地震
xqdispangle = []
yqdispangle = []
#顺风
xwpluswdispangle = []
xwminuswdispangle = []
ywpluswdispangle = []
ywminuswdispangle = []
#横风
xwhpluswdispangle = []
xwhminuswdispangle = []
ywhpluswdispangle = []
ywhminuswdispangle = []

#位移比
xplusdispratio = []
xplusdispratiod = []
xminusdispratio = []
xminusdispratiod = []

yplusdispratio = []
yplusdispratiod = []
yminusdispratio = []
yminusdispratiod = []

def readwmass():
    thefile = ".\设计结果\wmass.out"
    
    if not os.path.isfile(thefile):
        print("文件不存在，或者没有计算过。")
    else:
        with open(thefile) as file_object:
            for line in file_object:
                # ~ #地下室层数
                # ~ if line.endswith('设计参数输出\n'):
                    # ~ for j in range(6):
                        # ~ next(file_object)
                    # ~ basenumber = next(file_object).split()[1]
                    
                #总的楼层组装数
                if line.endswith('楼层属性\n'):
                    for j in range(3):
                        next(file_object)
                    sumfloor = int(next(file_object).split()[0])
                    
                #刚度比
                if line.endswith('各层刚心、偏心率、相邻层侧移刚度比等计算信息\n'):
                    for j in range(19):
                        next(file_object)
                    rat1 = next(file_object)
                    ratx1.append(rat1.split()[1])
                    raty1.append(rat1.split()[3])
                    rat2 = next(file_object)
                    ratx2.append(rat2.split()[1])
                    raty2.append(rat2.split()[3])
                    
                    for t in range(sumfloor - 1):
                        for i in range(9):
                            next(file_object)
                        rat1 = next(file_object)
                        ratx1.append(rat1.split()[1])
                        raty1.append(rat1.split()[3])
                        rat2 = next(file_object)
                        ratx2.append(rat2.split()[1])
                        raty2.append(rat2.split()[3])
    return sumfloor

def createexcel(sumfloor):
    wb = openpyxl.Workbook()
    
    #层间刚度比_框架
    sheetrat1 = wb.create_sheet(index = 0, title = '层间刚度比_框架')
    sheetrat1['A1'] = '层数'
    sheetrat1['B1'] = 'Ratx1'
    sheetrat1['C1'] = 'Raty1'
    for row in range(2, sumfloor + 2):
        sheetrat1.cell(column = 1, row = row, value = row - 1)
        sheetrat1.cell(column = 2, row = row, value = float(ratx1[row - 2]))
        sheetrat1.cell(column = 3, row = row, value = float(raty1[row - 2]))
    
    #层间刚度比_框支
    sheetrat2 = wb.create_sheet(index = 1, title = '层间刚度比_框支')
    sheetrat2['A1'] = '层数'
    sheetrat2['B1'] = 'Ratx2'
    sheetrat2['C1'] = 'Raty2'
    for row in range(2, sumfloor + 2):
        sheetrat2.cell(column = 1, row = row, value = row - 1)
        sheetrat2.cell(column = 2, row = row, value = float(ratx2[row - 2]))
        sheetrat2.cell(column = 3, row = row, value = float(raty2[row - 2]))
    
    #层间位移_分数
    sheetdriftfraction = wb.create_sheet(index = 2, title = '层间位移_分数')
    sheetdriftfraction['A1'] = '层数'
    sheetdriftfraction['B1'] = 'X 方向地震作用下的楼层最大位移'
    sheetdriftfraction['C1'] = 'Y 方向地震作用下的楼层最大位移'
    for row in range(2, sumfloor + 2):
        sheetdriftfraction.cell(column = 1, row = row, value = sumfloor + 2 - row)
        sheetdriftfraction.cell(column = 2, row = row, value = xqdispangle[row - 2])
        sheetdriftfraction.cell(column = 3, row = row, value = yqdispangle[row - 2])
        
    #层间位移_小数
    sheetdriftdecimal = wb.create_sheet(index = 3, title = '层间位移_小数')
    sheetdriftdecimal['A1'] = '层数'
    sheetdriftdecimal['B1'] = 'X 方向地震作用下的楼层最大位移'
    sheetdriftdecimal['C1'] = 'Y 方向地震作用下的楼层最大位移'
    for row in range(2, sumfloor + 2):
        sheetdriftdecimal.cell(column = 1, row = row, value = sumfloor + 2 - row)
        # 把字符串转换为分数
        # https://stackoverflow.com/questions/1806278/convert-fraction-to-float
        xfloatnumber = float(sum(Fraction(s) for s in xqdispangle[row - 2].split()))
        sheetdriftdecimal.cell(column = 2, row = row, value = xfloatnumber)
        yfloatnumber = float(sum(Fraction(s) for s in yqdispangle[row - 2].split()))
        sheetdriftdecimal.cell(column = 3, row = row, value = yfloatnumber)
    
    #位移比
    sheetdriftratio = wb.create_sheet(index = 4, title = '层间位移比')
    sheetdriftratio.merge_cells('B1:C1')
    sheetdriftratio['B1'] = 'X+ 偶然偏心规定水平力作用下的楼层最大位移'
    sheetdriftratio.merge_cells('D1:E1')
    sheetdriftratio['D1'] = 'X- 偶然偏心规定水平力作用下的楼层最大位移'
    sheetdriftratio.merge_cells('F1:G1')
    sheetdriftratio['F1'] = 'Y+ 偶然偏心规定水平力作用下的楼层最大位移'
    sheetdriftratio.merge_cells('H1:I1')
    sheetdriftratio['H1'] = 'Y- 偶然偏心规定水平力作用下的楼层最大位移'
    
    sheetdriftratio['A2'] = '层数'
    sheetdriftratio['B2'] = '最大位移'
    sheetdriftratio['C2'] = '层间位移'
    sheetdriftratio['D2'] = '最大位移'
    sheetdriftratio['E2'] = '层间位移'
    sheetdriftratio['F2'] = '最大位移'
    sheetdriftratio['G2'] = '层间位移'
    sheetdriftratio['H2'] = '最大位移'
    sheetdriftratio['I2'] = '层间位移'
    
    for row in range(3, sumfloor + 3):
        sheetdriftratio.cell(column = 1, row = row, value = sumfloor + 3 - row)
        sheetdriftratio.cell(column = 2, row = row, value = float(xplusdispratio[row - 3]))
        sheetdriftratio.cell(column = 3, row = row, value = float(xplusdispratiod[row - 3]))
        sheetdriftratio.cell(column = 4, row = row, value = float(xminusdispratio[row - 3]))
        sheetdriftratio.cell(column = 5, row = row, value = float(xminusdispratiod[row - 3]))
        sheetdriftratio.cell(column = 6, row = row, value = float(yplusdispratio[row - 3]))
        sheetdriftratio.cell(column = 7, row = row, value = float(yplusdispratiod[row - 3]))
        sheetdriftratio.cell(column = 8, row = row, value = float(yminusdispratio[row - 3]))
        sheetdriftratio.cell(column = 9, row = row, value = float(yminusdispratiod[row - 3]))
             
    wb.save('盈建科软件后处理程序v1.0.xlsx')


def readwdisp(sumfloor): 
    thefile = ".\设计结果\wdisp.out"
    
    if not os.path.isfile(thefile):
        print("文件不存在，或者没有计算过。")
    else:
        with open(thefile) as file_object:
            # stackoverflow: Using python, how to read a file starting at the seventh line ?
            for line in file_object:

                if line.endswith('X 方向地震作用下的楼层最大位移\n'):
                    # x向地震作用位移角
                    for j in range(4):
                        next(file_object)
                    for i in range(sumfloor):
                        next(file_object)
                        stringtemp = next(file_object)[49:55]
                        # 删除中间的空格 ''.join(x.split())
                        nostringtemp = ''.join(stringtemp.split())
                        # 把分数形式的字符串传给列表
                        xqdispangle.append(nostringtemp)

                    #print(xqdispangle)
                elif line.endswith('Y 方向地震作用下的楼层最大位移\n'):
                    # y向地震作用位移角
                    for j in range(4):
                        next(file_object)
                    for i in range(sumfloor):
                        next(file_object)
                        stringtemp = next(file_object)[49:55]
                        # 删除中间的空格 ''.join(x.split())
                        nostringtemp = ''.join(stringtemp.split())
                        yqdispangle.append(nostringtemp)
                        
                elif line.endswith('X+ 偶然偏心规定水平力作用下的楼层最大位移\n'):
                    # Ratio-(X) 最大位移比；Ratio-Dx 层间位移比
                    for j in range(4):
                        next(file_object)
                    for i in range(sumfloor):
                        xplusdispratio.append(next(file_object)[50:54])
                        xplusdispratiod.append(next(file_object)[50:54])
                    
                elif line.endswith('X- 偶然偏心规定水平力作用下的楼层最大位移\n'):
                    for j in range(4):
                        next(file_object)
                    for i in range(sumfloor):
                        xminusdispratio.append(next(file_object)[50:54])
                        xminusdispratiod.append(next(file_object)[50:54])
                        
                elif line.endswith('Y+ 偶然偏心规定水平力作用下的楼层最大位移\n'):
                    for j in range(4):
                        next(file_object)
                    for i in range(sumfloor):
                        yplusdispratio.append(next(file_object)[50:54])
                        yplusdispratiod.append(next(file_object)[50:54])
                        
                elif line.endswith('Y- 偶然偏心规定水平力作用下的楼层最大位移\n'):
                    for j in range(4):
                        next(file_object)
                    for i in range(sumfloor):
                        yminusdispratio.append(next(file_object)[50:54])
                        yminusdispratiod.append(next(file_object)[50:54])

def yjkpostprocessor():
    sumfloor = readwmass()
    readwdisp(sumfloor)
    createexcel(sumfloor)
    
yjkpostprocessor()        
